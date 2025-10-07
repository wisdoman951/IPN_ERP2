from __future__ import annotations

import argparse
import datetime as dt
import decimal
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook

IDENTITY_SHEETS: List[Tuple[str, str]] = [
    ("直營店", "DIRECT_STORE"),
    ("加盟店", "FRANCHISE"),
    ("合夥商", "PARTNER"),
    ("推廣商(分店能量師)", "PROMOTER"),
    ("B2B合作專案", "B2B_PROJECT"),
    ("心耀商", "XIN_YAO_MERCHANT"),
    ("會員", "MEMBER"),
    ("一般售價", "GENERAL_RETAIL"),
]

IDENTITY_PRIORITY: Dict[str, int] = {
    "DIRECT_STORE": 10,
    "FRANCHISE": 20,
    "PARTNER": 30,
    "PROMOTER": 40,
    "B2B_PROJECT": 50,
    "XIN_YAO_MERCHANT": 60,
    "MEMBER": 70,
    "GENERAL_RETAIL": 90,
}

PRICE_BOOK_NAME: Dict[str, str] = {
    "DIRECT_STORE": "直營店商品價目表",
    "FRANCHISE": "加盟店商品價目表",
    "PARTNER": "合夥商商品價目表",
    "PROMOTER": "推廣商商品價目表",
    "B2B_PROJECT": "B2B合作專案商品價目表",
    "XIN_YAO_MERCHANT": "心耀商商品價目表",
    "MEMBER": "會員商品價目表",
    "GENERAL_RETAIL": "一般售價商品價目表",
}

PRICE_DECIMAL = decimal.Decimal('0.00')


def _ensure_decimal(value: object) -> decimal.Decimal:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError("價格欄位不可為空")
    if isinstance(value, decimal.Decimal):
        return value.quantize(PRICE_DECIMAL)
    if isinstance(value, (int, float)):
        return decimal.Decimal(str(value)).quantize(PRICE_DECIMAL)
    if isinstance(value, str):
        return decimal.Decimal(value.strip()).quantize(PRICE_DECIMAL)
    raise TypeError(f"Unsupported price type: {type(value)!r}")


def normalize_product_name(raw_name: str) -> str:
    name = raw_name.strip()
    if not name:
        raise ValueError("產品名稱不可為空")
    name = name.replace('抺', '抹')
    name = name.replace('--', '-')
    # remove identity specific annotations
    for token in [
        '一般售價', '會員價', '豐盛領導', '推廣商', '能量師', '合夥商', '心耀商', 'B2B合作專案',
        '加盟店', '直營店', '價目表', '價目', '員工價', '一般售', '會員', '豐盛', '推廣', '能量', '合夥',
    ]:
        name = name.replace(token, '')
    # remove stray characters
    name = name.replace('價', '')
    name = name.replace('  ', ' ')
    name = re.sub(r'\s+', '', name)
    # normalize punctuation spacing
    name = name.replace('／', '/')
    name = name.replace('–', '-')
    return name


@dataclass
class ProductRecord:
    canonical_code: str
    base_name: str
    base_price: decimal.Decimal
    first_seen_identity: str


@dataclass
class PriceRecord:
    identity: str
    canonical_key: str
    custom_code: str
    custom_name: str
    price: decimal.Decimal


def read_price_workbook(path: Path) -> Tuple[Dict[str, ProductRecord], List[PriceRecord]]:
    workbook = load_workbook(filename=path, data_only=True)
    canonical_products: Dict[str, ProductRecord] = {}
    price_records: List[PriceRecord] = []

    for sheet_name, identity in IDENTITY_SHEETS:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"找不到工作表 '{sheet_name}'，請確認檔案格式")
        sheet = workbook[sheet_name]
        headers = None
        for row in sheet.iter_rows(values_only=True):
            if headers is None:
                headers = [str(cell).strip() if cell is not None else '' for cell in row]
                continue
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            row_map = {headers[i]: row[i] for i in range(len(headers))}
            code = str(row_map.get('產品代碼') or row_map.get('Product Code') or '').strip()
            name = str(row_map.get('產品名稱') or row_map.get('Name') or '').strip()
            price_value = row_map.get('售價') or row_map.get('Price')
            if not code or not name:
                raise ValueError(f"工作表 {sheet_name} 存在缺少產品代碼或名稱的資料列: {row_map}")
            price = _ensure_decimal(price_value)
            normalized_name = normalize_product_name(name)
            product = canonical_products.get(normalized_name)
            if product is None:
                product = ProductRecord(
                    canonical_code=code,
                    base_name=normalized_name,
                    base_price=price,
                    first_seen_identity=identity,
                )
                canonical_products[normalized_name] = product
            elif identity == 'GENERAL_RETAIL':
                product = ProductRecord(
                    canonical_code=product.canonical_code,
                    base_name=product.base_name,
                    base_price=price,
                    first_seen_identity=product.first_seen_identity,
                )
                canonical_products[normalized_name] = product
            price_records.append(
                PriceRecord(
                    identity=identity,
                    canonical_key=normalized_name,
                    custom_code=code,
                    custom_name=name.replace('抺', '抹').strip(),
                    price=price,
                )
            )
    return canonical_products, price_records


def sql_escape(value: str) -> str:
    return value.replace("'", "''")


def build_sql(canonical: Dict[str, ProductRecord], prices: Iterable[PriceRecord], valid_from: dt.date) -> str:
    lines: List[str] = []
    lines.append('-- generated by import_member_pricebooks.py')
    lines.append('START TRANSACTION;')

    for product in canonical.values():
        lines.append(
            "INSERT INTO product (code, name, price, status) "
            f"VALUES ('{sql_escape(product.canonical_code)}', '{sql_escape(product.base_name)}', {product.base_price}, 'PUBLISHED') "
            "ON DUPLICATE KEY UPDATE name = VALUES(name), price = VALUES(price), status = 'PUBLISHED';"
        )

    for identity, price_book_name in PRICE_BOOK_NAME.items():
        priority = IDENTITY_PRIORITY.get(identity, 100)
        lines.append(
            "INSERT INTO member_price_book (name, identity_type, scope_type, status, priority, valid_from) "
            f"VALUES ('{sql_escape(price_book_name)}', '{identity}', 'PRODUCT', 'ACTIVE', {priority}, '{valid_from.isoformat()}') "
            "ON DUPLICATE KEY UPDATE status = VALUES(status), priority = VALUES(priority), valid_from = VALUES(valid_from);"
        )
        lines.append(
            "DELETE mpi FROM member_price_book_item mpi "
            "JOIN member_price_book mpb ON mpi.price_book_id = mpb.price_book_id "
            f"WHERE mpb.name = '{sql_escape(price_book_name)}' AND mpb.identity_type = '{identity}' "
            "AND mpi.item_type = 'PRODUCT';"
        )

    for record in prices:
        product = canonical[record.canonical_key]
        price_book_name = PRICE_BOOK_NAME[record.identity]
        lines.append(
            "INSERT INTO member_price_book_item (price_book_id, item_type, item_id, price, currency, min_quantity, max_quantity, custom_code, custom_name, metadata, status) "
            "SELECT mpb.price_book_id, 'PRODUCT', p.product_id, {price}, 'TWD', 1, NULL, '{code}', '{name}', NULL, 'ACTIVE' "
            "FROM member_price_book mpb "
            "JOIN product p ON p.code = '{product_code}' "
            f"WHERE mpb.name = '{sql_escape(price_book_name)}' AND mpb.identity_type = '{record.identity}' "
            "ON DUPLICATE KEY UPDATE price = VALUES(price), custom_name = VALUES(custom_name), status = 'ACTIVE';"
            .format(
                price=record.price,
                code=sql_escape(record.custom_code),
                name=sql_escape(record.custom_name),
                product_code=sql_escape(product.canonical_code),
            )
        )

    lines.append('COMMIT;')
    return '\n'.join(lines) + '\n'


def main() -> None:
    parser = argparse.ArgumentParser(description='從會員別售價 Excel 匯出 SQL 腳本')
    parser.add_argument('excel_path', type=Path, help='會員別售價.xlsx 檔案路徑')
    parser.add_argument('--valid-from', type=str, default=dt.date.today().isoformat(), help='價目表生效日期 (YYYY-MM-DD)')
    parser.add_argument('--output', type=Path, help='輸出 SQL 檔案路徑，未指定時輸出至標準輸出')
    parser.add_argument('--dump-json', type=Path, help='額外輸出中介資料為 JSON 方便檢查 (選填)')
    args = parser.parse_args()

    valid_from = dt.date.fromisoformat(args.valid_from)

    canonical, price_records = read_price_workbook(args.excel_path)
    sql_script = build_sql(canonical, price_records, valid_from)

    if args.dump_json:
        payload = {
            'generated_at': dt.datetime.utcnow().isoformat() + 'Z',
            'valid_from': valid_from.isoformat(),
            'canonical_products': [
                {
                    'canonical_code': record.canonical_code,
                    'base_name': record.base_name,
                    'base_price': str(record.base_price),
                    'first_seen_identity': record.first_seen_identity,
                }
                for record in canonical.values()
            ],
            'price_records': [
                {
                    'identity': record.identity,
                    'canonical_key': record.canonical_key,
                    'custom_code': record.custom_code,
                    'custom_name': record.custom_name,
                    'price': str(record.price),
                }
                for record in price_records
            ],
        }
        args.dump_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2))

    if args.output:
        args.output.write_text(sql_script, encoding='utf-8')
    else:
        print(sql_script)


if __name__ == '__main__':
    main()
