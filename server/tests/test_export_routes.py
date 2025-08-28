import io
import os
import sys
import pytest
from datetime import date
from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app import create_app

@pytest.fixture
def client():
    app = create_app()
    app.config['TESTING'] = True
    with app.test_client() as client:
        yield client

def auth_headers():
    return {
        'X-Store-ID': '1',
        'X-Store-Level': 'admin'
    }

def test_member_export(client, monkeypatch):
    sample = [{
        'member_id': 1,
        'member_code': 'M001',
        'name': 'Alice',
        'birthday': '1990-01-01',
        'address': 'addr',
        'phone': '123',
        'gender': '女',
        'blood_type': 'A',
        'line_id': 'line',
        'inferrer_id': None,
        'occupation': 'engineer',
        'note': '',
        'store_id': 1
    }]
    monkeypatch.setattr('app.routes.member.get_all_members', lambda store_level, store_id: sample)
    rv = client.get('/api/member/export', headers=auth_headers())
    assert rv.status_code == 200
    assert rv.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def test_stress_test_export(client, monkeypatch):
    sample = [{
        'ipn_stress_id': 1,
        'member_code': 'M001',
        'Name': 'Alice',
        'position': '職員',
        'a_score': 1,
        'b_score': 2,
        'c_score': 3,
        'd_score': 4,
        'total_score': 10,
        'test_date': '2024-01-01'
    }]
    monkeypatch.setattr('app.routes.stress_test.get_all_stress_tests', lambda level, store_id, filters: sample)
    rv = client.get('/api/stress-test/export', headers=auth_headers())
    assert rv.status_code == 200
    assert rv.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def test_therapy_record_export(client, monkeypatch):
    sample = [{
        'therapy_record_id': 1,
        'member_code': 'M001',
        'member_name': 'Alice',
        'store_name': 'Store',
        'staff_name': 'Bob',
        'date': date(2024, 1, 1),
        'note': '',
        'deduct_sessions': 1,
        'remaining_sessions': 9
    }]
    monkeypatch.setattr('app.routes.therapy.export_therapy_records', lambda store_id: sample)
    rv = client.get('/api/therapy/record/export', headers=auth_headers())
    assert rv.status_code == 200
    assert rv.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def test_therapy_record_export_respects_store(client, monkeypatch):
    called = {}

    def fake_export(store_id):
        called['store_id'] = store_id
        return []

    monkeypatch.setattr('app.routes.therapy.export_therapy_records', fake_export)
    monkeypatch.setattr('app.routes.therapy.get_user_from_token',
                        lambda req: {'store_id': 2, 'permission': 'staff'})

    rv = client.get('/api/therapy/record/export', headers=auth_headers())
    assert rv.status_code == 200
    assert called['store_id'] == 2


def test_therapy_record_export_empty(client, monkeypatch):
    monkeypatch.setattr('app.routes.therapy.export_therapy_records', lambda store_id: [])
    rv = client.get('/api/therapy/record/export', headers=auth_headers())
    assert rv.status_code == 200
    wb = load_workbook(filename=io.BytesIO(rv.data))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == ['療程記錄ID', '會員編號', '會員姓名', '商店名稱', '服務人員', '日期', '備註', '扣除堂數', '療程剩餘數']
    assert ws.max_row == 1

def test_sales_order_export(client, monkeypatch):
    sample = [{
        'order_id': 1,
        'order_number': 'SO1',
        'order_date': '2024-01-01',
        'grand_total': 100,
        'sale_category': 'P',
        'note': '',
        'member_name': 'Alice',
        'staff_name': 'Bob'
    }]
    monkeypatch.setattr('app.routes.sales_order_routes.get_all_sales_orders', lambda keyword=None: {'success': True, 'data': sample})
    rv = client.get('/api/sales-orders/export', headers=auth_headers())
    assert rv.status_code == 200
    assert rv.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
